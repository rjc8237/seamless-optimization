import json
import os
from pathlib import Path
import sys
import pandas as pd
import igl
import numpy as np
base_dir = os.path.dirname(__file__)
module_dir = os.path.join(base_dir, '..', 'py')
sys.path.append(module_dir)
script_dir = os.path.join(base_dir, '..', 'ext', 'penner-optimization', 'scripts')
sys.path.append(script_dir)
import symdir

OUTPUT_DIR = Path("./output")

def num_triangles_with_energy(X, E_min):
    count = sum(1 for e in X if e >= E_min)
    e_max = max(X)
    return count, e_max

def create_solver_comparison_table(output_dir, cg_lp=2, llt_lp=1, p=None, err=None):
    """Create a table comparing metrics (total_time, E_worst, etc.) for different solvers
    with side-by-side columns for CG and Ch_LLT, and a single faces column.
    """
    
    solvers_config = {"CG": cg_lp, "Ch_LLT": llt_lp}
    data = []
    
    output_dir = output_dir / "test13_s"

    # Get mesh folders from CG solver
    cg_base_path = output_dir / ("Lp_" + str(cg_lp) + "_" + str(p) + "_" + str(err)) / "CG"
    
    for mesh_folder in cg_base_path.iterdir():
        if not mesh_folder.is_dir():
            continue
        
        json_files = list(mesh_folder.glob("*.json"))
        if not json_files:
            continue
        
        # Extract mesh name from folder
        mesh_name = mesh_folder.name.replace("_output", "")
        row_data = {"Mesh": mesh_name}
        metrics_by_solver = {}
        faces_value = None
        thresholds = np.linspace(0.25, 2.5, 10)

        for solver, lp in solvers_config.items():
            metrics = {}
            try:
                solver_base_path = output_dir / ("Lp_" + str(lp) + "_" + str(p) + "_" + str(err)) / solver
                solver_mesh_folder = solver_base_path / mesh_folder.name
                
                if solver_mesh_folder.exists() and solver_mesh_folder.is_dir():
                    json_files = list(solver_mesh_folder.glob("*.json"))
                    if json_files:
                        with open(json_files[0], 'r') as f:
                            json_data = json.load(f)
                            metrics["total_time"] = json_data.get("total_time")
                            metrics["E_worst"] = json_data["opt_log"][-1].get("E_worst")
                            metrics["E_avg"] = json_data["opt_log"][-1].get("E_avg", "N/A")
                            metrics["converge_reason"] = json_data.get("converge_reason", "N/A")
                            faces_value = json_data["opt_log"][0].get("F_size", "N/A")
                            X = json_data.get("energy", [])
                            for threshold in thresholds:
                                key = f"E>={threshold}"
                                count, _ = num_triangles_with_energy(X, threshold)
                                metrics[key] = count / faces_value * 100.0

            except Exception as e:
                print(f"Error getting data for {mesh_name} - {solver} - Lp{lp}: {e}")
            
            metrics_by_solver[solver] = metrics

        metrics_list = ["total_time", "E_worst", "E_avg", "converge_reason"] + [f"E>={thresh}" for thresh in thresholds]
        for metric_name in metrics_list:
            for solver in ["CG", "Ch_LLT"]:
                col_key = f"{metric_name}_{solver}"
                value = metrics_by_solver.get(solver, {}).get(metric_name, "N/A")
                row_data[col_key] = f"{value:.4f}" if isinstance(value, float) else value
        
        row_data["faces"] = faces_value if faces_value is not None else "N/A"
        data.append(row_data)
    
    df = pd.DataFrame(data)
    
    column_order = ["Mesh", "faces"]
    for metric in metrics_list:
        for solver in ["CG", "Ch_LLT"]:
            column_order.append(f"{metric}_{solver}")
    df = df[column_order]
    
    # dynamic title using per-solver n values and m
    title = (f"Comparison CG(L{cg_lp}), Ch_LLT(L{llt_lp}), p = {p}, err = {err}")
    print("\n" + "="*200)
    print(title)
    print("="*200)
    print(df.to_string(index=False))
    print("="*200 + "\n")
    
    # Save to CSV with title above header
    out_fn = f"comparison_p={p}_err={err}.csv"
    output_file = output_dir / "statistics" / "tables" / out_fn
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, "w", newline="") as f:
        f.write(title + "\n\n")
        df.to_csv(f, index=False)
    
    print(f"Table saved to {output_file}")
    return df

def get_mesh_folders(output_dir, Lp_value, n, m, solver):
    mesh_folders = {}

    output_dir = output_dir / ("Lp_" + str(Lp_value) + "_" + str(n) + "_" + str(m))
    output_dir = output_dir / solver
    for folder in output_dir.iterdir():
        if not folder.is_dir():
            continue

        # Check if folder contains .obj files
        obj_files = list(folder.glob("*.obj"))
        if obj_files:
            obj_path = obj_files[0]
            obj_name = obj_path.stem.split("_refined_with_uv_out_")[0]

            # Count faces in the OBJ file
            face_count = 0
            with open(obj_path, 'r') as f:
                for line in f:
                    if line.startswith('f '):
                        face_count += 1
        
            # Only include meshes with ~100K faces (allowing ±10% tolerance)
            if 90000 <= face_count <= 110000:
                mesh_folders[obj_path] = obj_name
                print(f"  Added {folder.name} with {face_count} faces")

    return mesh_folders

def create_solver_lp_comparison_table(output_dir, lp_values=[1, 2, 3], solvers=["CG", "Ch_LLT"]):
    """Create a table comparing total_time for different solvers and Lp values
    with hierarchical headers: L_1, L_2, L_3 each with CG and Ch_LLT subheaders
    """
    
    # Get mesh folders for first Lp value to get mesh names
    mesh_folders = get_mesh_folders(output_dir, lp_values[0], solvers[0])
    
    if not mesh_folders:
        print(f"No mesh folders found")
        return None
    
    # Initialize data dictionary
    data = {}
    
    # Iterate through each mesh
    for obj_path, mesh_name in sorted(mesh_folders.items()):
        row_data = {}
        
        # For each Lp and solver combination, get total_time and E_worst
        for lp in lp_values:
            for solver in solvers:
                total_time = None
                e_worst = None
                
                try:
                    # Build path to solver folder
                    solver_base_path = output_dir / ("Lp_" + str(lp)) / solver
                    
                    # Find matching mesh folder
                    for mesh_folder in solver_base_path.iterdir():
                        if not mesh_folder.is_dir():
                            continue
                        
                        if not mesh_folder.name.endswith("_output"):
                            continue
                        
                        # Check if this is the right mesh
                        obj_files = list(mesh_folder.glob("*.obj"))
                        if obj_files:
                            obj_name = obj_files[0].stem.split("_refined_with_uv_out_")[0]
                            
                            if obj_name == mesh_name:
                                # Found the right mesh, get total_time and E_worst from JSON
                                json_files = list(mesh_folder.glob("*.json"))
                                if json_files:
                                    with open(json_files[0], 'r') as f:
                                        json_data = json.load(f)
                                        total_time = json_data.get("total_time")
                                        e_worst = json_data.get("E_worst")
                                break
                
                except Exception as e:
                    print(f"Error getting data for {mesh_name} - {solver} - Lp{lp}: {e}")
                
                # Create column name
                col_key = f"L{lp}_{solver}"
                
                # Format the cell value: total_time / E_worst (ratio)
                if total_time is not None and e_worst is not None:
                    cell_value = f"{total_time:.2f} / {e_worst:.2e}"
                else:
                    cell_value = "N/A"
                
                row_data[col_key] = cell_value
        
        data[mesh_name] = row_data
    
    # Create DataFrame from the data
    df_list = []
    for mesh_name in sorted(data.keys()):
        row = {"Mesh": mesh_name}
        row.update(data[mesh_name])
        df_list.append(row)
    
    df = pd.DataFrame(df_list)
    
    # Display table
    print("\n" + "="*200)
    print(f"Total Time / E_worst Comparison Table")
    print(f"Format: total_time / E_worst")
    print("="*200)
    print(df.to_string(index=False))
    print("="*200 + "\n")
        
    # Save with custom hierarchical headers to CSV
    output_file = output_dir / "statistics" / "tables" / f"comparison_table_time_per_energy.csv"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_file, 'w') as f:
        # Write first header row (Lp values)
        f.write("Mesh")
        for lp in lp_values:
            f.write(f",L_{lp},L_{lp}")
        f.write("\n")
        
        # Write second header row (solver names)
        f.write("Mesh")
        for lp in lp_values:
            for solver in solvers:
                f.write(f",{solver}")
        f.write("\n")
        
        # Write data rows
        for _, row in df.iterrows():
            f.write(f"{row['Mesh']}")
            for lp in lp_values:
                for solver in solvers:
                    col_key = f"L{lp}_{solver}"
                    f.write(f",{row[col_key]}")
            f.write("\n")
    
    print(f"Table saved to {output_file}")
    
    # Also create Excel version with merged headers
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Side, Font
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison"
        
        # Write header row 1: Mesh, L_1, L_1, L_2, L_2, L_3, L_3
        ws['A1'] = "Mesh"
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(bold=True)
        
        col = 2
        for lp in lp_values:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            cell = ws.cell(row=1, column=col)
            cell.value = f"L_{lp}"
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            col += 2
        
        # Write header row 2: CG, Ch_LLT for each Lp
        ws['A2'] = "Mesh"
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].font = Font(bold=True)
        
        col = 2
        for lp in lp_values:
            for solver in solvers:
                cell = ws.cell(row=2, column=col)
                cell.value = solver
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
                col += 1
        
        # Write data rows
        for idx, row in df.iterrows():
            ws.cell(row=idx+3, column=1).value = row["Mesh"]
            col = 2
            for lp in lp_values:
                for solver in solvers:
                    col_key = f"L{lp}_{solver}"
                    ws.cell(row=idx+3, column=col).value = row[col_key]
                    col += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(lp_values) * len(solvers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 35
        
        # Set row heights for headers
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25
        
        excel_file = output_dir / "statistics" / "tables" / f"comparison_table_time_per_energy.xlsx"
        wb.save(excel_file)
        print(f"Excel table with merged headers saved to {excel_file}")
    
    except ImportError:
        print("openpyxl not installed. Install with: pip install openpyxl")
    
    # Create ratio table for easier analysis
    ratio_data = {}
    for mesh_name in sorted(data.keys()):
        row_data = {}
        for lp in lp_values:
            for solver in solvers:
                col_key = f"L{lp}_{solver}"
                cell_value = data[mesh_name][col_key]
                
                # Extract ratio from cell_value
                if cell_value != "N/A":
                    try:
                        ratio_str = cell_value.split("(")[1].rstrip(")")
                        row_data[col_key] = float(ratio_str)
                    except:
                        row_data[col_key] = None
                else:
                    row_data[col_key] = None
        
        ratio_data[mesh_name] = row_data
    
    # Create ratio DataFrame
    ratio_list = []
    for mesh_name in sorted(ratio_data.keys()):
        row = {"Mesh": mesh_name}
        row.update(ratio_data[mesh_name])
        ratio_list.append(row)
    
    df_ratios = pd.DataFrame(ratio_list)
    
    print("\n" + "="*180)
    print(f"Ratio Table: total_time / E_worst")
    print("="*180)
    print(df_ratios.to_string(index=False))
    print("="*180 + "\n")
    
    # Save ratio table with custom headers
    ratio_file = output_dir / "statistics" / "tables" / f"ratio_table_time_per_energy.csv"
    with open(ratio_file, 'w') as f:
        # Write first header row (Lp values)
        f.write("Mesh")
        for lp in lp_values:
            f.write(f",L_{lp},L_{lp}")
        f.write("\n")
        
        # Write second header row (solver names)
        f.write("Mesh")
        for lp in lp_values:
            for solver in solvers:
                f.write(f",{solver}")
        f.write("\n")
        
        # Write data rows
        for _, row in df_ratios.iterrows():
            f.write(f"{row['Mesh']}")
            for lp in lp_values:
                for solver in solvers:
                    col_key = f"L{lp}_{solver}"
                    val = row[col_key]
                    if val is not None:
                        f.write(f",{val:.2f}")
                    else:
                        f.write(",N/A")
            f.write("\n")
    
    print(f"Ratio table saved to {ratio_file}")
    
    return df, df_ratios

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="Create comparison tables")
    parser.add_argument("--Lp", nargs="+", type=int, default=[1, 2, 3], help="Lp values")
    parser.add_argument("--solvers", nargs="+", default=["CG", "Ch_LLT"], help="Solvers to compare")
    
    args = parser.parse_args()

    output_dir = OUTPUT_DIR
    ps = [1, 3, 5, 8, 10]
    errs = [0.05, 0.1]
    for p in ps:
        for err in errs:
            create_solver_comparison_table(output_dir, 2, 1, p, err)

if __name__ == "__main__":
    main()